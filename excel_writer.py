import os
from datetime import date
from openpyxl import Workbook, load_workbook

def append_data_to_master(sheet_name, data_rows, platform_name):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(base_dir, "data")
    file_path = os.path.join(data_dir, "hackathons_master.xlsx")

    os.makedirs(data_dir, exist_ok=True)

    headers = ["Title", "Deadline", "Apply Link", "Platform", "Date Added"]
    new_entries_count = 0

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

    existing_links = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 3 and row[2]:
            existing_links.add(str(row[2]).strip())

    today_str = date.today().strftime("%Y-%m-%d")

    for title, deadline, link in data_rows:
        clean_link = str(link).strip()
        if clean_link and clean_link not in existing_links:
            ws.append([title, deadline, clean_link, platform_name, today_str])
            existing_links.add(clean_link)
            new_entries_count += 1

    wb.save(file_path)
    return new_entries_count
